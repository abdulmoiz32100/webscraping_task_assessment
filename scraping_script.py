import json
import time
import re
from seleniumbase import SB
from openpyxl import Workbook, load_workbook
import os


XLSX_FILE = "investorlift_listings.xlsx"
HEADERS = [
    "Listing Title", "Price", "ARV", "Property Address", "Bedrooms",
    "Bathrooms", "Square Footage", "Listing URL", "Status",
    "Probability Score", "Posted Date"
]

LOCATIONS = ["Milwaukee, WI, USA", "Columbus, OH, USA"]
ALLOWED_STATES = ["WI", "OH"]


def get_or_create_workbook():
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Listings"
        ws.append(HEADERS)
    return wb, ws


def is_allowed_state(address):
    for state in ALLOWED_STATES:
        if f", {state}," in address or f", {state} " in address or address.strip().endswith(f", {state}"):
            return True
    return False


def scrape_listings(sb):
    cards = sb.find_elements(".ui-deal-card_wrapper.ui-deal-card")
    listings = []

    for card in cards:
        try:
            try:
                price = card.find_element("css selector", ".ui-deal-card-property-price-and-arv_value").text.strip()
            except Exception:
                price = ""

            try:
                arv = card.find_element("css selector", ".ui-deal-card-property-price-and-arv_avg").text.strip()
                arv = arv.replace("(", "").replace(")", "").replace("ARV -", "").strip()
            except Exception:
                arv = ""

            try:
                address = card.find_element("css selector", ".ui-deal-card-property-address").text.strip()
            except Exception:
                address = ""

            if address and not is_allowed_state(address):
                print(f"[~] Skipping (not OH/WI): {address}")
                continue

            beds, baths, sqft = "", "", ""
            try:
                detail_divs = card.find_elements("css selector", ".ui-deal-card-property-details-row > div")
                for div in detail_divs:
                    text = div.text.strip()
                    if "Bed" in text:
                        beds = re.search(r"(\d+)", text)
                        beds = beds.group(1) if beds else text
                    elif "Bath" in text:
                        baths = re.search(r"(\d+)", text)
                        baths = baths.group(1) if baths else text
                    elif "sq" in text.lower():
                        sqft = text.replace("sq.ft", "").strip()
            except Exception:
                pass

            try:
                link = card.find_element("css selector", "a.ui-deal-card-link")
                url = link.get_attribute("href")
            except Exception:
                url = ""

            try:
                status = card.find_element("css selector", ".ui-deal-card-badge.state-for-sale").text.strip()
            except Exception:
                status = ""

            try:
                prob = card.find_element("css selector", ".ui-deal-card-overlay_probability-score").text.strip()
            except Exception:
                prob = ""

            try:
                posted = card.find_element("css selector", ".ui-message-status-badge").text.strip()
            except Exception:
                posted = ""

            listings.append({
                "Listing Title": address,
                "Price": price,
                "ARV": arv,
                "Property Address": address,
                "Bedrooms": beds,
                "Bathrooms": baths,
                "Square Footage": sqft,
                "Listing URL": url,
                "Status": status,
                "Probability Score": prob,
                "Posted Date": posted,
            })
        except Exception as e:
            print(f"[!] Error scraping a card: {e}")
            continue

    return listings


def save_to_xlsx(listings):
    wb, ws = get_or_create_workbook()
    for listing in listings:
        ws.append([listing[h] for h in HEADERS])
    wb.save(XLSX_FILE)
    print(f"[+] Saved {len(listings)} listings to {XLSX_FILE}")


def scrape_location(sb, location, all_urls_seen):
    print(f"\n{'='*50}")
    print(f"[*] Searching for: {location}")
    print(f"{'='*50}")

    search_input = '.ui-dropdown-autocomplete input[type="text"]'
    sb.wait_for_element(search_input, timeout=15)

    try:
        close_btn = '.ui-dropdown-autocomplete .ui-close-button'
        if sb.is_element_visible(close_btn):
            sb.click(close_btn)
            time.sleep(1)
    except Exception:
        pass

    sb.click(search_input)
    sb.type(search_input, location)
    time.sleep(2)

    try:
        sb.wait_for_element('.ui-dropdown-autocomplete .ui-dropdown_list', timeout=10)
        sb.click('.ui-dropdown-autocomplete .ui-dropdown_list div:first-child')
    except Exception:
        sb.send_keys(search_input, "\n")

    print("[*] Waiting for listings to load...")
    sb.sleep(5)
    sb.wait_for_element('.ui-deal-card_wrapper.ui-deal-card', timeout=20)
    sb.sleep(3)

    scroller = '.map-and-catalog-with-filters_content-and-filters_cards_scroller_inner'
    total_saved = 0

    while True:
        listings = scrape_listings(sb)

        new_listings = []
        for l in listings:
            if l["Listing URL"] not in all_urls_seen and l["Listing URL"]:
                all_urls_seen.add(l["Listing URL"])
                new_listings.append(l)

        if new_listings:
            save_to_xlsx(new_listings)
            total_saved += len(new_listings)
            print(f"[*] Total unique listings for {location}: {total_saved}")

        sb.execute_script(
            f"""
            var scroller = document.querySelector('{scroller}');
            if (scroller) {{
                scroller.parentElement.scrollTop += 1000;
            }} else {{
                window.scrollBy(0, 1000);
            }}
            """
        )
        sb.sleep(3)

        listings_after = scrape_listings(sb)
        new_urls = [l["Listing URL"] for l in listings_after if l["Listing URL"] not in all_urls_seen]

        if not new_urls:
            sb.execute_script(
                f"""
                var scroller = document.querySelector('{scroller}');
                if (scroller) {{
                    scroller.parentElement.scrollTop += 2000;
                }} else {{
                    window.scrollBy(0, 2000);
                }}
                """
            )
            sb.sleep(4)
            listings_after = scrape_listings(sb)
            new_urls = [l["Listing URL"] for l in listings_after if l["Listing URL"] not in all_urls_seen]

            if not new_urls:
                print(f"[*] No more new listings for {location}. Moving on.")
                break

    return total_saved


def login_and_scrape():
    with SB(uc=True, headless=False) as sb:
        sb.uc_open_with_reconnect("https://investorlift.com/marketplace/", 4)

        sb.wait_for_element('button.ui-button span.content:contains("Log in")', timeout=15)
        sb.click('button.ui-button span.content:contains("Log in")')

        sb.wait_for_element('.enter-email-form input[type="text"]', timeout=15)
        sb.type('.enter-email-form input[type="text"]', "")

        sb.click('.enter-email-form button[type="submit"]')

        sb.wait_for_element('.enter-password-form input[type="password"]', timeout=15)
        time.sleep(1)

        sb.type('.enter-password-form input[type="password"]', "")

        sb.js_click('.enter-password-form .ui-checkbox input[type="checkbox"]')

        sb.click('.enter-password-form button[type="submit"]')

        sb.sleep(5)

        cookies = sb.get_cookies()
        with open("cookies.json", "w") as f:
            json.dump(cookies, f, indent=2)
        print(f"\n[+] {len(cookies)} cookies saved to cookies.json")

        all_urls_seen = set()
        grand_total = 0

        for location in LOCATIONS:
            count = scrape_location(sb, location, all_urls_seen)
            grand_total += count

        print(f"\n{'='*50}")
        print(f"[+] ALL DONE! Total unique listings scraped: {grand_total}")
        print(f"[+] Data saved to {XLSX_FILE}")
        print(f"{'='*50}")


if __name__ == "__main__":
    login_and_scrape()